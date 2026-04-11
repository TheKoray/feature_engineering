from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
from datetime import date
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
import plotly.io as pio

from override import (getRatingData, getReasonData, getBolgeData, getOverallData, getAllModelsRatingData)

app = Flask(__name__)

SEGMENT_LIST = [
    'TÜM MODELLER', '8-40-ÜRETİM', '8-40-HİZMET',
    '40+-ÜRETİM', '40+-HİZMET', '40+-TİCARET', '40+-İNŞAAT'
]

def today():
    return date.today().strftime('%d.%m.%Y')

# ---------------------------------------------------------------------------
# SAYFA ROUTE'LARI
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template('index.html', run_date=today())

@app.route('/musteri')
def musteri():
    return render_template('musteri.html', segments=SEGMENT_LIST, run_date=today())

@app.route('/grup')
def grup():
    return render_template('grup.html', segments=SEGMENT_LIST, run_date=today())


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

@app.route('/api/data')
def api_data():
    view  = request.args.get('view',  'musteri')
    tab   = request.args.get('tab',   'Rating')
    model = request.args.get('model', 'TÜM MODELLER')

    musteri_or_grup = 'MUSTERI' if view == 'musteri' else 'GRUP'

    if tab == 'Rating':
        # TÜM MODELLER seçildiğinde getAllModelsRatingData() kullan (segment kırılımı yok)
        # Diğer segmentler için getRatingData() kullan
        if model == 'TÜM MODELLER':
            df = getAllModelsRatingData(musteri_or_grup=musteri_or_grup)
        else:
            df = getRatingData(musteri_or_grup=musteri_or_grup, model_tipi=model)

        # Kolon isimleri: -5,-4,-3,-2,-1, RATING, RATING_ADET, OVERRIDE_ADET, +0,+1,+2,+3,+4,+5
        neg_cols = ['-5', '-4', '-3', '-2', '-1']
        pos_cols = ['+0', '+1', '+2', '+3', '+4', '+5']
        bar_cols = neg_cols + pos_cols

        # TOTAL satırını ayır
        total_mask = df['RATING'].astype(str).str.upper() == 'TOTAL'
        df_plot    = df[~total_mask].copy()
        df_total   = df[total_mask].copy()
        df_out     = pd.concat([df_plot, df_total], ignore_index=True)

        # OVERRIDE_ORAN zaten SQL'den geliyor, formatlama yapılmıyor (zaten DECIMAL olarak hesaplanmış)

        # ── GRAFİK ──────────────────────────────────────────────
        fig = make_subplots(specs=[[{'secondary_y': True}]])

        # Bar: OVERRIDE_ADET (text olarak OVERRIDE_ORAN göster)
        override_text = (df_plot['OVERRIDE_ORAN'].astype(str) + '%').values
        fig.add_trace(
            go.Bar(
                x=df_plot['RATING'].astype(str),
                y=df_plot['OVERRIDE_ADET'],
                name='Override Adet',
                marker_color='#3b82f6',
                text=override_text,
                textposition='outside',
            ),
            secondary_y=False
        )

        # Line: OVERRIDE_ADET (trend)
        fig.add_trace(
            go.Scatter(
                x=df_plot['RATING'].astype(str),
                y=df_plot['OVERRIDE_ADET'],
                name='Trend',
                mode='lines+markers',
                line=dict(color='red', width=1.5),
                marker=dict(size=5),
            ),
            secondary_y=True
        )

        fig.update_layout(
            title=dict(text='Override Adet Dağılımı', x=0.5,
                       font=dict(size=18, color='#ffffff')),
            template='plotly_white',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
                        font=dict(color='#94a3b8')),
                        width = 1200,
            margin=dict(t=80, r=20, b=60, l=60),
        )
        fig.update_xaxes(title_text='Rating', type='category',
                         title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))
        fig.update_yaxes(title_text='Adet', secondary_y=False,
                         title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))
        fig.update_yaxes(title_text='Trend', secondary_y=True, showgrid=False,
                         title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))

        # ── TABLO ────────────────────────────────────────────────
        # Gösterilecek kolonlar (OVERRIDE_ORAN da dahil)
        show_cols = neg_cols + ['RATING', 'RATING_ADET', 'OVERRIDE_ADET', 'OVERRIDE_ORAN'] + pos_cols

        # Sadece mevcut kolonları al
        show_cols = [c for c in show_cols if c in df_out.columns]

        table_data    = df_out[show_cols].to_dict(orient='records')
        table_columns = show_cols

        return jsonify({
            'table':   table_data,
            'columns': table_columns,
            'figure':  json.loads(fig.to_json()),
        })

    if tab == 'Reasons':
        import plotly.express as px
        df = getReasonData(musteri_or_grup=musteri_or_grup, model_tipi=model)
        df.columns = [c.strip().replace('"', '') for c in df.columns]
        df = df.rename(columns={'-': 'NEGATIF', '+': 'POZITIF'})

        total_mask = df['OVERRIDE_REASON'].astype(str).str.upper() == 'TOTAL'
        df_plot    = df[~total_mask].copy()

        fig = px.bar(
            df_plot,
            x='OVERRIDE_REASON',
            y=['NEGATIF', 'POZITIF'],
            barmode='group',
            color_discrete_map={'NEGATIF': '#dc2626', 'POZITIF': '#16a34a'},
            text_auto=True,
        )
        fig.update_traces(textfont_size=11, textposition='outside', cliponaxis=False)
        fig.update_layout(
            title=dict(text='Override Reasons', x=0.5, font=dict(size=18, color='#ffffff')),
            width=1200,
            xaxis_title='Override Reason',
            yaxis_title='Adet',
            template='plotly_white',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
                        font=dict(color='#94a3b8')),
        )
        fig.update_xaxes(title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))
        fig.update_yaxes(title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))
        fig.update_xaxes(type='category')

        return jsonify({
            'table':   df.to_dict(orient='records'),
            'columns': list(df.columns),
            'figure':  json.loads(fig.to_json()),
        })

    if tab == 'Bölge':
        import plotly.express as px
        df = getBolgeData(musteri_or_grup=musteri_or_grup, model_tipi=model)
        df.columns = [c.strip() for c in df.columns]

        # Grafik için float değerleri tutacak DataFrame (tablo için orijinal kalsın)
        df_chart = df.copy()

        # Yüzdelik değerleri sayıya çevir (string "61.3%" → float 61.3) - GRAFİK İÇİN
        for col in ['NEGATIF_OVERRIDE_ORAN', 'POZITIF_OVERRIDE_ORAN', 'OVERRIDE_ORAN']:
            if col in df_chart.columns:
                df_chart[col] = df_chart[col].astype(str).str.replace('%', '').astype(float)

        fig = px.bar(
            df_chart,
            x='BOLGE',
            y=['NEGATIF_OVERRIDE_ORAN', 'POZITIF_OVERRIDE_ORAN', 'OVERRIDE_ORAN'],
            barmode='group',
            color_discrete_map={
                'NEGATIF_OVERRIDE_ORAN': '#dc2626',
                'POZITIF_OVERRIDE_ORAN': '#16a34a',
                'OVERRIDE_ORAN':         '#3b82f6',
            },
            text_auto=True,
        )
        fig.update_traces(textfont_size=11, textposition='outside', cliponaxis=False)
        fig.update_layout(
            title=dict(text='Bölgelere Göre Override Dağılımı', x=0.5, font=dict(size=18, color='#ffffff')),
            width=1200,
            xaxis_title='Bölge',
            yaxis_title='Oran (%)',
            template='plotly_white',
            paper_bgcolor='#1e293b',
            plot_bgcolor='#0f172a',
            font=dict(color='#94a3b8'),
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
                        font=dict(color='#94a3b8')),
            margin=dict(t=80, r=20, b=60, l=60),
        )
        fig.update_xaxes(title_text='Bölge', title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'), type='category')
        fig.update_yaxes(title_text='Oran (%)', title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))

        return jsonify({
            'table':   df.to_dict(orient='records'),  # Orijinal (% ile birlikte)
            'columns': list(df.columns),
            'figure':  json.loads(fig.to_json()),  # df_chart'tan grafik (float)
        })

    if tab == 'Overall':
        import plotly.express as px
        df = getOverallData(musteri_or_grup=musteri_or_grup, model_tipi=model)
        if df.empty:
            return jsonify({'table': [], 'columns': [], 'figure': {}, 'msg': 'Veri bulunamadı.'})

        fig = px.bar(
            df,
            x=df.columns[0],
            y=['NEGATIF_ADET', 'POZITIF_ADET'],
            barmode='group',
            color_discrete_map={'NEGATIF_ADET': '#dc2626', 'POZITIF_ADET': '#16a34a'},
            text_auto=True,
        )
        fig.update_traces(textfont_size=11, textposition='outside', cliponaxis=False)
        fig.update_layout(
            title=dict(text='Genel Bakış', x=0.5, font=dict(size=18, color='#ffffff')),
            width=1200,
            xaxis_title='Özet',
            yaxis_title='Adet',
            template='plotly_white',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1,
                        font=dict(color='#94a3b8')),
        )
        fig.update_xaxes(title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))
        fig.update_yaxes(title_font=dict(color='#ffffff'), tickfont=dict(color='#94a3b8'))
        fig.update_xaxes(type='category')

        return jsonify({
            'table':   df.to_dict(orient='records'),
            'columns': list(df.columns),
            'figure':  json.loads(fig.to_json()),
        })

    return jsonify({'table': [], 'columns': [], 'figure': {}})


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------

def rgb_to_hex(r, g, b, intensity):
    """RGB'yi hex'e çevir ve intensity'ye göre karıştır (beyaz ile blend)"""
    # intensity: 0..1 (0 açık=beyaz, 1 koyu=tam renk)
    # RGB değerlerini intensity'ye göre karıştır
    r_final = int(255 - (255 - r) * intensity)
    g_final = int(255 - (255 - g) * intensity)
    b_final = int(255 - (255 - b) * intensity)

    return f"{r_final:02X}{g_final:02X}{b_final:02X}"


def get_scale_color(col_name, value, max_val):
    """Renk scale'i hesapla"""
    SCALE_COLS = {
        'NEGATIF_OVERRIDE_ORAN': (220, 38, 38),
        'POZITIF_OVERRIDE_ORAN': (22, 163, 74),
        'NEGATIF': (220, 38, 38),
        'POZITIF': (22, 163, 74),
    }

    if col_name not in SCALE_COLS or max_val == 0:
        return None

    try:
        val_str = str(value).strip().replace('%', '')
        val = float(val_str)
        r, g, b = SCALE_COLS[col_name]
        intensity = min(val / max_val, 1.0)  # 0..1 (capped at 1)

        return rgb_to_hex(r, g, b, intensity)
    except Exception as e:
        print(f"Color calc error for {col_name}={value}: {e}")
        return None


def get_max_values(table_data, columns):
    """Her kolon için max değeri hesapla (TOTAL satırları hariç)"""
    maxes = {}
    scale_cols = {'NEGATIF_OVERRIDE_ORAN', 'POZITIF_OVERRIDE_ORAN', 'NEGATIF', 'POZITIF'}

    for col in columns:
        if col in scale_cols:
            vals = []
            for row in table_data:
                # TOTAL satırını atla
                if any(str(v).strip().upper() == 'TOTAL' for v in row.values()):
                    continue

                try:
                    val_str = str(row.get(col, '0')).strip().replace('%', '')
                    val = float(val_str)
                    if val >= 0:
                        vals.append(val)
                except (ValueError, TypeError):
                    pass

            if vals:
                maxes[col] = max(vals)

    return maxes


@app.route('/export/excel', methods=['POST'])
def export_excel():
    """Tablo ve grafiği Excel'e kaydet"""
    try:
        data = request.json
        view = data.get('view', 'musteri')
        tab = data.get('tab', 'Rating')
        model = data.get('model', 'TÜM MODELLER')
        table_data = data.get('table', [])
        columns = data.get('columns', [])
        figure_json = data.get('figure', {})

        # Excel workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = tab

        # Tablo başlığı
        ws['A1'] = f"{tab} Tablosu - {model}"
        ws['A1'].font = Font(bold=True, size=12)

        # Max değerleri hesapla (renk scale için)
        maxes = get_max_values(table_data, columns)

        # Kolon başlıkları
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")

        # Tablo verileri (renk scale'i ile)
        for row_idx, row_data in enumerate(table_data, 4):
            # TOTAL satırı mı kontrol et
            is_total = False
            for v in row_data.values():
                if str(v).strip().upper() == 'TOTAL':
                    is_total = True
                    break

            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # TOTAL satırı değilse renk scale'i uygula
                if not is_total and col_name in maxes and maxes[col_name] > 0:
                    color_hex = get_scale_color(col_name, value, maxes[col_name])
                    if color_hex:
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)

        # Kolon genişliği
        for col_idx in range(1, len(columns) + 1):
            col_letter = chr(64 + col_idx) if col_idx < 27 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
            ws.column_dimensions[col_letter].width = 15

        # Grafik resmi (PNG olarak Plotly'den)
        chart_row = len(table_data) + 8
        if figure_json and figure_json.get('data') and len(figure_json.get('data', [])) > 0:
            try:
                import plotly.graph_objects as go
                fig = go.Figure(figure_json)

                # Grafik dosyası oluştur
                chart_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_chart.png')
                fig.write_image(chart_path, width=1000, height=600)

                # Excel'e grafik ekle
                if os.path.exists(chart_path):
                    img = XLImage(chart_path)
                    img.width = 650
                    img.height = 400
                    ws.add_image(img, f'A{chart_row}')
                    os.remove(chart_path)
            except Exception as e:
                print(f"⚠️ Grafik ekleme hatası: {e}")

        # Excel dosyasını kaydet
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{tab}_{model}_{timestamp}.xlsx"
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

        wb.save(filepath)
        print(f"✅ Excel kaydedildi: {filepath}")

        return jsonify({
            'success': True,
            'message': f'Excel dosyası kaydedildi: {filename}',
            'filename': filename
        })

    except Exception as e:
        print(f"❌ Export hatası: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Hata: {str(e)}'
        }), 500


# ---------------------------------------------------------------------------

if __name__ == '__main__':
    app.run(port=8888, debug=True)


